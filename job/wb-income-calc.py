import os
from collections import defaultdict
from openpyxl import load_workbook

# Путь к папке с файлами Excel
folder_path = "/Users/user/Documents/wb-income"

# Словарь для хранения количества товаров по артикулу для каждого файла
files_data = defaultdict(lambda: defaultdict(int))
# Словарь для хранения общего количества товаров по артикулу
summary_data = defaultdict(int)

# Перебираем все файлы в папке
for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx'):
        # Загружаем файл Excel
        wb = load_workbook(os.path.join(folder_path, filename))
        # Получаем активный лист
        sheet = wb.active
        # Пропускаем первую строку с заголовками
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Получаем артикул товара и количество
            article = row[3]  # Предполагаем, что артикул находится в четвертой колонке (нумерация с 0)
            quantity = 1  # Предполагаем, что количество находится в шестой колонке (нумерация с 0)
            # Добавляем количество товара в словарь для текущего файла и артикула
            files_data[filename][article] += quantity
            # Добавляем количество товара в общий словарь
            summary_data[article] += quantity

# Записываем результаты в файл txt
with open('/Users/user/Documents/result.txt', 'w') as file:
    # Перебираем данные из словаря
    for filename, data in files_data.items():
        # Записываем имя файла
        file.write(f'{filename}:\n')
        print(f'{filename}:\n')
        # Записываем количество товаров по артикулам
        for article, quantity in data.items():
            file.write(f'Артикул: {article}, Количество: {quantity}\n')
            print(f'Артикул: {article}, Количество: {quantity}\n')
        # Разделяем данные для разных файлов пустой строкой
        file.write('\n')

    # Добавляем секцию SUMMARY
    file.write('SUMMARY:\n')
    print('SUMMARY:\n')
    # Записываем общее количество товаров по артикулам
    for article, quantity in summary_data.items():
        file.write(f'Артикул: {article}, Общее количество: {quantity}\n')
        print(f'Артикул: {article}, Общее количество: {quantity}\n')
